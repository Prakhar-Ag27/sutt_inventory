from curses import is_term_resized
from imp import reload
import random
from django.shortcuts import render
import os
from .models import Item, Through, Student
from django.http import HttpResponse
from django.shortcuts import render 
import datetime
from django.utils import timezone
import smtplib
from openpyxl import Workbook, load_workbook


def details(request):
    item_list = Item.objects.all()
    context = {'item_list': item_list, 'type' : "issue"}
    if request.user.is_authenticated:
        return render(request, 'fnapp/detailspage.html', context)
        
    else:
        return render(request, 'fnapp/notlogin.html')

def issue(request, itemcode):
    tobeissued = Item.objects.get(unique_code=itemcode)
    context = {'tobeissued': tobeissued}
    if request.user.is_authenticated:
        return render(request, 'fnapp/issuepage.html', context)
    else:
        return render(request, 'fnapp/notlogin.html')

def return_items(request):
    item_list = Through.objects.filter(student__email_id = request.user.email).order_by('-time_of_return')
    context = {'item_list': item_list, 'type' : "return"}
    if request.user.is_authenticated:
        return render(request, 'fnapp/detailspage.html', context)
    else:
        return render(request, 'fnapp/notlogin.html')

def success_issue(request, itemcode):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            pass
        else:
            obj = Through()
            obj.item = Item.objects.get(unique_code=itemcode)
            obj.student = Student.objects.get(email_id=request.user.email)
            obj.qty_issued = int(request.POST['items'])
            obj.current_issued = int(request.POST['items'])
            req_time = datetime.datetime.now()
            obj.time_of_issue = req_time
            obj.time_of_issue_int = int(req_time.strftime("%Y%m%d%H%M%S"))
            obj.item.total_qty = obj.item.total_qty- int(request.POST['items'])
            obj.item.current_out = obj.item.current_out + int(request.POST['items'])
            obj.save()
            obj.item.save()

            with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()
                smtp.login("agrawal.prakhar35@gmail.com", "tndumdixwmkppymx")
                subject = 'Regarding issue of item from IMS'
                body = f'You have successfully issued {obj.item.name_of_item} with item code {obj.item.unique_code}.'
                msg = f'Subject: {subject}\n\n{body}'
                smtp.sendmail("agrawal.prakhar35@gmail.com", request.user.email, msg)

                return render(request, 'fnapp/successpage_issue.html', {"itemcode" : itemcode})

    else:
        return render(request, 'fnapp/notlogin.html')

def success_return(request, itemcode, temp, timeofissueint):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            pass
        else:
            obj = Through.objects.filter(item__unique_code = itemcode).get(time_of_issue_int = timeofissueint)
            if temp == 1:
                obj.qty_returned = obj.current_issued
                obj.item.total_qty = obj.item.total_qty + obj.qty_returned
                obj.item.current_out = obj.item.current_out - obj.qty_returned
                obj.time_of_return = timezone.now()
                obj.current_issued = 0
                obj.save()
                obj.item.save()

                with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
                    smtp.ehlo()
                    smtp.starttls()
                    smtp.ehlo()
                    smtp.login("agrawal.prakhar35@gmail.com", "tndumdixwmkppymx")
                    subject = 'Regarding return of issued item from IMS'
                    body = f'You have successfully returned {obj.item.name_of_item} with item code {obj.item.unique_code}.'
                    msg = f'Subject: {subject}\n\n{body}'
                    smtp.sendmail("agrawal.prakhar35@gmail.com", request.user.email, msg)

                    return render(request, 'fnapp/successpage_return.html', {"itemcode" : itemcode})

            if temp == 0:
                obj.qty_returned = int(request.POST['items'])
                obj.item.total_qty = obj.item.total_qty + obj.qty_returned
                obj.item.current_out = obj.item.current_out - obj.qty_returned
                obj.time_of_return = timezone.now()
                obj.current_issued = obj.current_issued - int(request.POST['items'])
                obj.save()
                obj.item.save()

                with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
                    smtp.ehlo()
                    smtp.starttls()
                    smtp.ehlo()
                    smtp.login("agrawal.prakhar35@gmail.com", "tndumdixwmkppymx")
                    subject = 'Regarding issue of item from IMS'
                    body = f'You have successfully returned {obj.qty_returned} out of {obj.qty_issued} issued {obj.item.name_of_item} with item code {obj.item.unique_code}'
                    msg = f'Subject: {subject}\n\n{body}'
                    smtp.sendmail("agrawal.prakhar35@gmail.com", request.user.email, msg)

                    return render(request, 'fnapp/successpage_return.html', {"itemcode" : itemcode})

    else:
        return render(request, 'fnapp/notlogin.html')

def log_add(request):
    logitems = Item.objects.all().order_by('-added_on')
    context = {'logitems': logitems}
    if request.user.is_authenticated:
        return render(request, 'fnapp/log_add.html', context)
    else:
        return render(request, 'fnapp/notlogin.html')

def log_issue(request):
    logitems = Through.objects.filter(qty_issued__gt = 0).order_by('-time_of_issue')
    context = {'logitems': logitems}
    if request.user.is_authenticated:
        return render(request, 'fnapp/log_issue.html', context)
    else:
        return render(request, 'fnapp/notlogin.html')

def log_return(request):
    logitems = Through.objects.filter(qty_returned__gt = 0).order_by('-time_of_return')
    context = {'logitems': logitems}
    if request.user.is_authenticated:
        return render(request, 'fnapp/log_return.html', context)
    else:
        return render(request, 'fnapp/notlogin.html')

def edit_item(request):
    item_list = Item.objects.all()
    context = {'item_list' : item_list, "type" : "edit"}
    return render(request, 'fnapp/detailspage.html', context)

def edit_item_two(request, itemcode):
    item_req = Item.objects.get(unique_code = itemcode)
    context = {"item_req": item_req, "itemcode": itemcode}
    return render(request, 'fnapp/edititemstwo.html', context)
    
def edit_item_three(request, itemcode):
    item_req = Item.objects.get(unique_code = itemcode)
    context = {'item_req' : item_req, "itemcode": itemcode}
    item_req.type_of_item = request.POST["type"]
    item_req.name_of_item = request.POST["name"]
    item_req.total_qty = int(request.POST["quantity"])
    item_req.current_out = int(request.POST["out"])
    item_req.save()
    return render(request, 'fnapp/successpage_edit.html', context)

def filter(request):
    name = request.POST['item']
    type = request.POST['temp']
    item_list = Item.objects.filter(name_of_item__iexact = name)
    if item_list:
        context = {'item_list': item_list, 'type' : type}
    else:
        item_list = Item.objects.filter(type_of_item__iexact = name)
        context = {'item_list': item_list, 'type' : type}
    return render(request, 'fnapp/detailspage.html', context)

def simple_upload(request):
    return render(request, 'fnapp/uploadpage.html')

def success_upload(request):
    wb = load_workbook(request.FILES['myfile'])
    ws = wb.active
    row = 2
    col = 0
    req_list = []
    while True:
        
        while True:
            char = chr(65 + col)
            cell = char + str(row)
            if ws[cell].value or ws[cell].value == 0:
                req_list.append(ws[cell].value)
                col = col + 1
            else:
                col = col + 1
                break;
        Item.objects.create(name_of_item = req_list[0], type_of_item = req_list[1], total_qty = int(req_list[2]), current_out = int(req_list[3]), unique_code = random.randint(10_000_000, 99_999_999))
        row = row + 1
        col = 0
        char = chr(65 + col)
        cell = char + str(row)
        req_list.clear()
        if ws[cell].value:
            pass
        else:
            break;

    return render(request, 'fnapp/successpage_upload.html')